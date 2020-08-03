#!/usr/bin/env python
# -*- coding: utf-8 -*-
import json
import pandas as pd
import re
from openpyxl import load_workbook


def parse_tolgas_b():
    url = "https://www.tolgas.ru/abitur/bachelor/#abitur_vstupitelnye-ispytania"
    table_vi = pd.read_html(url)[1]
    table_vi.to_excel("src/excel/tolgas_vi.xlsx")
    table_pp = pd.read_html(url)[0]
    table_pp.to_excel("src/excel/tolgas_pp.xlsx")
    updateExcel_tolgas_b()


def updateExcel_tolgas_b():
    wb_vi = load_workbook("src/excel/tolgas_vi.xlsx")
    wb_pp = load_workbook("src/excel/tolgas_pp.xlsx")

    sheet_ranges_vi = wb_vi['Sheet1']
    column_b_vi = sheet_ranges_vi['B']
    column_k_vi = sheet_ranges_vi['K']
    column_j_vi = sheet_ranges_vi['J']
    column_l_vi = sheet_ranges_vi['L']
    column_m_vi = sheet_ranges_vi['M']
    column_n_vi = sheet_ranges_vi['N']
    column_o_vi = sheet_ranges_vi['O']
    column_p_vi = sheet_ranges_vi['P']
    column_q_vi = sheet_ranges_vi['Q']

    sheet_ranges_pp = wb_pp['Sheet1']
    column_b_pp = sheet_ranges_pp['B']
    column_d_pp = sheet_ranges_pp['D']
    column_e_pp = sheet_ranges_pp['E']
    column_f_pp = sheet_ranges_pp['F']
    column_g_pp = sheet_ranges_pp['G']
    column_h_pp = sheet_ranges_pp['H']
    column_i_pp = sheet_ranges_pp['I']
    column_j_pp = sheet_ranges_pp['J']
    column_k_pp = sheet_ranges_pp['K']

    for i in range(4, len(column_b_vi)):
        if column_b_vi[i].value is None:
            pass
        elif re.match(r"\d\d\.", column_b_vi[i].value):
            codeAndProgram = column_b_vi[i].value.split(" ", 1)
            column_j_vi[i].value = codeAndProgram[1]
        else:
            pass
    wb_vi.save("src/excel/tolgas_vi.xlsx")

    for n in range(len(column_b_vi)):
        for m in range(len(column_b_pp)):
            if column_j_vi[n].value == column_d_pp[m].value:
                column_k_vi[n].value = column_e_pp[m].value
                column_l_vi[n].value = column_f_pp[m].value
                column_m_vi[n].value = column_g_pp[m].value
                column_n_vi[n].value = column_h_pp[m].value
                column_o_vi[n].value = column_i_pp[m].value
                column_p_vi[n].value = column_j_pp[m].value
                column_q_vi[n].value = column_k_pp[m].value
    wb_vi.save("src/excel/tolgas_vi.xlsx")
    for u in range(len(column_b_vi)):
        if column_k_vi[u].value is None:
            column_k_vi[u].value = '0'
            column_l_vi[u].value = '0'
            column_m_vi[u].value = '0'
            column_n_vi[u].value = '0'
            column_o_vi[u].value = '0'
            column_p_vi[u].value = '0'
            column_q_vi[u].value = '0'
    wb_vi.save("src/excel/tolgas_vi.xlsx")


def viewExcel_tolgas_b():
    parse_tolgas_b()
    wb_vi = load_workbook("src/excel/tolgas_vi.xlsx")

    sheet_ranges_vi = wb_vi['Sheet1']
    column_b_vi = sheet_ranges_vi['B']
    column_c_vi = sheet_ranges_vi['C']
    column_e_vi = sheet_ranges_vi['E']
    column_g_vi = sheet_ranges_vi['G']
    column_k_vi = sheet_ranges_vi['K']
    column_l_vi = sheet_ranges_vi['L']
    column_m_vi = sheet_ranges_vi['M']
    column_n_vi = sheet_ranges_vi['N']
    column_o_vi = sheet_ranges_vi['O']
    column_p_vi = sheet_ranges_vi['P']
    column_q_vi = sheet_ranges_vi['Q']

    out_all = []
    for i in range(1, len(column_b_vi)):
        if column_b_vi[i].value is None:
            pass
        elif re.match(r"\d\d\.", column_b_vi[i].value):
            codeAndProgram = column_b_vi[i].value.split(" ", 1)
            pay_o = column_p_vi[i].value
            pay_z = column_q_vi[i].value
            kcp = int(column_k_vi[i].value) - (int(pay_o.replace('-', '0')) + int(pay_z.replace('-', '0')))
            out_all.append({
                'code': str(codeAndProgram[0]),
                'program': str(codeAndProgram[1]),
                'level': 'bachelor',
                'subject_1': str(column_c_vi[i - 1].value),
                'ball_1': str(column_c_vi[i].value),
                'subject_2': str(column_e_vi[i - 1].value),
                'ball_2': str(column_e_vi[i].value),
                'subject_3': str(column_g_vi[i - 1].value),
                'ball_3': str(column_g_vi[i].value),
                'subject_4': "-",
                'ball_4': "-",
                'plan_all': str(column_k_vi[i].value),
                'kcp': str(kcp),
                'special_o': str(column_l_vi[i].value),
                'special_z': str(column_m_vi[i].value),
                'special_oz': "-",
                'general_o': str(column_n_vi[i].value),
                'general_z': str(column_o_vi[i].value),
                'general_oz': "-",
                'goal_o': "-",
                'goal_oz': "-",
                'goal_z': "-",
                'pay_o': str(column_p_vi[i].value),
                'pay_z': str(column_q_vi[i].value),
                'pay_oz': "-"
            })
    return out_all


with open('src/tolgas_bach.json', 'w', encoding="utf-8") as fp:
    json.dump(viewExcel_tolgas_b(), fp, ensure_ascii=False)